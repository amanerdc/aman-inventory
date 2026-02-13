from __future__ import annotations

import csv
from datetime import date, datetime
from typing import Iterable, Sequence


def export_to_csv(path: str, columns: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)


def export_to_excel(
    path: str,
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    image_paths: Sequence[str | None] | None = None,
    image_height: int = 60,
    header_lines: Sequence[str] | None = None,
    image_column: int = 0,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.drawing.image import Image as XLImage  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
        from openpyxl.styles import Font  # type: ignore

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        columns_out = list(columns)
        if image_paths is not None:
            columns_out.insert(max(0, min(image_column, len(columns_out))), "Picture")
        header_offset = 0
        if header_lines:
            for line in header_lines:
                ws.append([line])
            for row_idx in range(1, len(header_lines) + 1):
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
            header_offset = len(header_lines)
        ws.append(columns_out)
        row_values = list(rows)
        for row in row_values:
            if image_paths is not None:
                row_out = list(row)
                row_out.insert(max(0, min(image_column, len(row_out))), "")
                ws.append(row_out)
            else:
                ws.append(list(row))
        for i in range(1, len(columns_out) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        if image_paths is not None:
            photo_col = max(0, min(image_column, len(columns_out) - 1))
            photo_letter = get_column_letter(photo_col + 1)
            ws.column_dimensions[photo_letter].width = 16
            for idx, img_path in enumerate(image_paths, start=2 + header_offset):
                if not img_path:
                    continue
                try:
                    xl_img = XLImage(img_path)
                    if xl_img.height:
                        scale = image_height / xl_img.height
                        xl_img.height = image_height
                        xl_img.width = int(xl_img.width * scale)
                    else:
                        xl_img.height = image_height
                        xl_img.width = image_height
                    ws.add_image(xl_img, f"{photo_letter}{idx}")
                    ws.row_dimensions[idx].height = image_height + 6
                except Exception:
                    continue
        wb.save(path)
    except Exception:
        # Fallback to CSV if openpyxl is not installed
        csv_path = path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if header_lines:
                for line in header_lines:
                    writer.writerow([line])
                writer.writerow([])
            if image_paths is not None:
                columns_out = list(columns)
                columns_out.insert(max(0, min(image_column, len(columns_out))), "Picture")
                writer.writerow(columns_out)
                row_values = list(rows)
                for idx, row in enumerate(row_values):
                    img_path = image_paths[idx] if idx < len(image_paths) else ""
                    row_out = list(row)
                    row_out.insert(max(0, min(image_column, len(row_out))), img_path or "")
                    writer.writerow(row_out)
            else:
                writer.writerow(list(columns))
                for row in rows:
                    writer.writerow(list(row))


def _simple_pdf(lines: Sequence[str]) -> bytes:
    # Minimal PDF with one page and a single Helvetica font.
    content = []
    y = 800
    for line in lines:
        safe = line.replace("(", "[").replace(")", "]")
        content.append(f"1 0 0 1 50 {y} Tm ({safe}) Tj")
        y -= 14
    stream = "\n".join(content)

    objects = []
    objects.append("1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj")
    objects.append("2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj")
    objects.append(
        "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj"
    )
    objects.append("4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj")
    objects.append(f"5 0 obj << /Length {len(stream.encode('utf-8'))} >> stream\n{stream}\nendstream endobj")

    xref_positions = []
    pdf = ["%PDF-1.4"]
    for obj in objects:
        xref_positions.append(sum(len(p) + 1 for p in pdf))
        pdf.append(obj)
    xref_start = sum(len(p) + 1 for p in pdf)

    xref = ["xref", f"0 {len(objects) + 1}", "0000000000 65535 f "]
    for pos in xref_positions:
        xref.append(f"{pos:010} 00000 n ")

    trailer = [
        "trailer",
        f"<< /Size {len(objects) + 1} /Root 1 0 R >>",
        "startxref",
        str(xref_start),
        "%%EOF",
    ]

    full_pdf = "\n".join(pdf + xref + trailer)
    return full_pdf.encode("utf-8")


def export_to_pdf(
    path: str,
    title: str,
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    image_paths: Sequence[str | None] | None = None,
    image_height: int = 60,
    header_lines: Sequence[str] | None = None,
    image_column: int = 0,
) -> None:
    try:
        from reportlab.lib.pagesizes import letter  # type: ignore
        from reportlab.lib.utils import ImageReader  # type: ignore
        from reportlab.pdfgen import canvas  # type: ignore

        c = canvas.Canvas(path, pagesize=letter)
        width, height = letter
        y = height - 50
        if header_lines:
            c.setFont("Helvetica-Bold", 12)
            for line in header_lines:
                c.drawString(50, y, line)
                y -= 16
            y -= 8
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, title)
        y -= 24
        c.setFont("Helvetica", 10)
        columns_out = list(columns)
        if image_paths is not None:
            columns_out.insert(max(0, min(image_column, len(columns_out))), "Picture")
        c.drawString(50, y, " | ".join(columns_out))
        y -= 18

        row_values = list(rows)
        for idx, row in enumerate(row_values):
            if image_paths is not None:
                img_path = image_paths[idx] if idx < len(image_paths) else None
                if img_path:
                    try:
                        img = ImageReader(img_path)
                        c.drawImage(img, 50, y - image_height + 10, height=image_height, preserveAspectRatio=True, mask="auto")
                    except Exception:
                        pass
                row_out = list(row)
                row_out.insert(max(0, min(image_column, len(row_out))), "")
                line = " | ".join(str(x) for x in row_out)
                c.drawString(50 + image_height + 10, y, line[:1500])
                y -= max(14, image_height + 6)
            else:
                line = " | ".join(str(x) for x in row)
                c.drawString(50, y, line[:1500])
                y -= 14
            if y < 50:
                c.showPage()
                y = height - 50
                c.setFont("Helvetica", 10)
        c.save()
    except Exception:
        lines = []
        if header_lines:
            lines.extend(list(header_lines))
            lines.append("")
        lines.extend([title, ""])
        columns_out = list(columns)
        if image_paths is not None:
            columns_out.insert(max(0, min(image_column, len(columns_out))), "Picture")
        lines.append(" | ".join(columns_out))
        row_values = list(rows)
        for idx, row in enumerate(row_values):
            if image_paths is not None:
                prefix = "[image]" if idx < len(image_paths) and image_paths[idx] else ""
                row_out = list(row)
                row_out.insert(max(0, min(image_column, len(row_out))), prefix)
                lines.append(" | ".join(str(x) for x in row_out))
            else:
                lines.append(" | ".join(str(x) for x in row))
        with open(path, "wb") as f:
            f.write(_simple_pdf(lines))


def export_to_jpg(
    path: str,
    title: str,
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    image_paths: Sequence[str | None] | None = None,
    image_height: int = 60,
    header_lines: Sequence[str] | None = None,
    image_column: int = 0,
) -> None:
    try:
        from PIL import Image, ImageDraw, ImageFont  # type: ignore

        row_values = list(rows)
        columns_out = list(columns)
        if image_paths is not None:
            columns_out.insert(max(0, min(image_column, len(columns_out))), "Picture")
        width = 1200
        line_height = max(20, image_height + 10) if image_paths is not None else 20
        extra_lines = 0
        if header_lines:
            extra_lines = len(header_lines) + 1
        height = max(200, (len(row_values) + 4 + extra_lines) * line_height)
        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        y = 10
        if header_lines:
            for line in header_lines:
                draw.text((10, y), line, fill="black", font=font)
                y += line_height
            y += 4
        draw.text((10, y), title, fill="black", font=font)
        y += line_height
        draw.text((10, y), " | ".join(columns_out), fill="black", font=font)
        y += line_height

        for idx, row in enumerate(row_values):
            x = 10
            if image_paths is not None:
                img_path = image_paths[idx] if idx < len(image_paths) else None
                if img_path:
                    try:
                        thumb = Image.open(img_path)
                        thumb.thumbnail((image_height, image_height))
                        img.paste(thumb, (x, y))
                    except Exception:
                        pass
                x += image_height + 10
                row_out = list(row)
                row_out.insert(max(0, min(image_column, len(row_out))), "")
                draw.text((x, y), " | ".join(str(x) for x in row_out), fill="black", font=font)
            else:
                draw.text((x, y), " | ".join(str(x) for x in row), fill="black", font=font)
            y += line_height
        img.save(path, "JPEG")
    except Exception:
        raise RuntimeError("JPG export requires Pillow. Install with: pip install pillow")


def export_airbnb_inspection_pdf(path: str, report_date: date, room_no: str, items: Sequence[dict]) -> None:
    title = "Airbnb Inspection Checklist"
    note = (
        "A refundable \u20b1500 deposit is required upon check-in to ensure all items on the inspection list "
        "remain in good condition. The deposit will be refunded based on the condition of the listed items upon checkout."
    )

    def _grouped_items() -> list[tuple[str, dict]]:
        grouped: list[tuple[str, dict]] = []
        last_area = None
        for item in items:
            area = item.get("brand") or ""
            if area != last_area:
                grouped.append((area, {}))
                last_area = area
            grouped.append(("", item))
        return grouped

    try:
        from reportlab.lib.pagesizes import letter, landscape  # type: ignore
        from reportlab.pdfgen import canvas  # type: ignore

        c = canvas.Canvas(path, pagesize=landscape(letter))
        page_width, page_height = landscape(letter)
        margin = 28
        gap = 24
        copy_width = (page_width - (2 * margin) - gap) / 2
        start_y = page_height - margin

        import textwrap

        def draw_copy(x_offset: float) -> None:
            y = start_y
            c.setFont("Helvetica-Bold", 12)
            c.drawString(x_offset, y, title)
            y -= 16
            c.setFont("Helvetica", 10)
            c.drawString(x_offset, y, f"Date: {report_date.strftime('%Y-%m-%d')}")
            c.drawString(x_offset + copy_width / 2, y, f"Room No.: {room_no}")
            y -= 18

            col_widths = [0.28, 0.42, 0.1, 0.2]
            col_x = [x_offset]
            for w in col_widths[:-1]:
                col_x.append(col_x[-1] + (copy_width * w))

            c.setFont("Helvetica-Bold", 10)
            c.drawString(col_x[0], y, "Area")
            c.drawString(col_x[1], y, "Item")
            c.drawString(col_x[2], y, "Qty")
            c.drawString(col_x[3], y, "Turn-over")
            y -= 14

            c.setFont("Helvetica", 9)
            for area, item in _grouped_items():
                if area:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(col_x[0], y, area)
                    y -= 12
                    c.setFont("Helvetica", 9)
                    continue
                if not item:
                    continue
                name = str(item.get("name") or "")
                qty = str(item.get("quantity") or "")
                c.drawString(col_x[1], y, name)
                c.drawString(col_x[2], y, qty)
                y -= 12

            y -= 6
            c.setFont("Helvetica", 9)
            c.drawString(x_offset, y, "Received by:")
            y -= 12
            c.drawString(x_offset, y, "Name & Signature")
            y -= 16
            c.drawString(x_offset, y, "Turn-over accepted by:")
            y -= 12
            c.drawString(x_offset, y, "Name & Signature")
            y -= 18

            c.setFont("Helvetica", 8)
            for line in textwrap.wrap(note, width=90):
                c.drawString(x_offset, y, line)
                y -= 10

        draw_copy(margin)
        draw_copy(margin + copy_width + gap)
        c.save()
    except Exception:
        base_lines = [
            title,
            f"Date: {report_date.strftime('%Y-%m-%d')}",
            f"Room No.: {room_no}",
            "",
            "Area | Item | Qty | Turn-over",
        ]
        for item in items:
            area = item.get("brand") or ""
            name = item.get("name") or ""
            qty = item.get("quantity") or ""
            base_lines.append(f"{area} | {name} | {qty} | ")
        base_lines.extend(
            [
                "",
                "Received by:",
                "Name & Signature",
                "",
                "Turn-over accepted by:",
                "Name & Signature",
                "",
                note,
            ]
        )
        lines = list(base_lines)
        lines.append("")
        lines.append("---- COPY 2 ----")
        lines.extend(base_lines)
        with open(path, "wb") as f:
            f.write(_simple_pdf(lines))
